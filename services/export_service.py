"""
services/export_service.py
Generates CSV and Excel exports for admin complaint downloads.
"""
import csv
import io
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

HEADERS    = ["Complaint ID", "Citizen Name", "Email", "Title", "Category",
              "Location", "Priority", "Status", "Department", "Date Submitted", "Last Updated"]
COL_WIDTHS = [22, 18, 28, 28, 20, 32, 10, 14, 18, 20, 20]

STATUS_FILL   = {"Pending": "FEF3C7", "In Progress": "DBEAFE", "Resolved":  "D1FAE5"}
PRIORITY_FILL = {"High":    "FEE2E2", "Medium":      "FEF9C3", "Low":       "DCFCE7"}


def _row_values(d: dict) -> list:
    created = d.get("created_at")
    updated = d.get("updated_at")
    return [
        d.get("complaint_id", ""), d.get("user_name", ""),  d.get("user_email", ""),
        d.get("title", ""),        d.get("category", ""),   d.get("location", ""),
        d.get("priority", ""),     d.get("status", ""),     d.get("department", ""),
        created.strftime("%Y-%m-%d %H:%M") if isinstance(created, datetime) else "",
        updated.strftime("%Y-%m-%d %H:%M") if isinstance(updated, datetime) else "",
    ]


def generate_csv(docs: list) -> str:
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(HEADERS)
    for d in docs:
        writer.writerow(_row_values(d))
    return out.getvalue()


def generate_excel(docs: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Complaints"

    thin        = Side(border_style="thin", color="D1D5DB")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill    = PatternFill("solid", fgColor="1A2E6B")
    hdr_font    = Font(bold=True, color="FFFFFF", size=11)
    hdr_align   = Alignment(horizontal="center", vertical="center")

    # Header row
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = cell_border
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, d in enumerate(docs, 2):
        status   = d.get("status",   "")
        priority = d.get("priority", "")
        for col_idx, val in enumerate(_row_values(d), 1):
            cell            = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border     = cell_border
            cell.alignment  = Alignment(vertical="center", wrap_text=(col_idx == 6))
            if col_idx == 8 and status in STATUS_FILL:
                cell.fill = PatternFill("solid", fgColor=STATUS_FILL[status])
                cell.font = Font(bold=True)
            if col_idx == 7 and priority in PRIORITY_FILL:
                cell.fill = PatternFill("solid", fgColor=PRIORITY_FILL[priority])
                cell.font = Font(bold=True)
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
